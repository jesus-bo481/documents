"""
Auto-GD — Fase 1: Llenar el Excel de cálculo con datos del proyecto
"""

import openpyxl
from datetime import datetime
from pathlib import Path


# ─────────────────────────────────────────────────────────────────
# 1. INPUTS DEL PROYECTO
# ─────────────────────────────────────────────────────────────────
PROYECTO = {
    "nombre":       "LA MARTINA 1",
    "ciudad":       "PARATEBUENO",
    "departamento": "CUNDINAMARCA",
    "fecha":        datetime(2026, 3, 31),
    "temp_min":     18,
    "temp_max":     30,
    "temp_prom":    28,
    # Longitudes de cable AC por inversor [m] (hoja MEMORIA, celdas AB24-AB26)
    # Tramo: salida del inversor → gabinete de protecciones SSFV
    # AB33 (totalizador → transformador) se deja como está en el Excel base
    "long_ac_inv": [54, 40, 28],
}

# Ruta al Excel de metrado (I1S1, I2S1, …)
EXCEL_METRADO = (
    r"C:\Users\JesúsAndrésBustilloO\Documents\GD\Proyectos"
    r"\MARTINA 1 Y 2\MARTINA_1\04. Editables\03. Tablas de Cálculo"
    r"\Calculo de metros cable MARTINA 1 - copia.xlsx"
)

# Ruta al Excel de cálculo base (con fórmulas)
EXCEL_BASE = (
    r"C:\Users\JesúsAndrésBustilloO\Documents\GD\Proyectos"
    r"\MARTINA 1 Y 2\MARTINA_1\04. Editables\03. Tablas de Cálculo"
    r"\PRGD-ENTRENAMIENTO CLAUDE.xlsm"
)

# Ruta de salida (copia del Excel base ya llenado)
EXCEL_SALIDA = (
    r"C:\Users\JesúsAndrésBustilloO\Documents\Auto-GD\output"
    r"\PRGD-CALCULO-{nombre}-{fecha}.xlsm"
).format(
    nombre=PROYECTO["nombre"].replace(" ", "_"),
    fecha=PROYECTO["fecha"].strftime("%Y%m%d"),
)


# ─────────────────────────────────────────────────────────────────
# 2. LEER METRADO
# ─────────────────────────────────────────────────────────────────
def leer_metrado(path: str) -> dict:
    """
    Lee el Excel de metrado y devuelve un dict:
    { "I1S1": (pos_m, neg_m), "I1S2": (pos_m, neg_m), ... }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    datos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        string_id = str(row[0]).strip()   # ej: I1S1
        pos_m     = float(row[1] or 0)
        neg_m     = float(row[2] or 0)
        datos[string_id] = (pos_m, neg_m)
    return datos


# ─────────────────────────────────────────────────────────────────
# 3. MAPEAR STRINGS DEL METRADO A FILAS DEL EXCEL DE CÁLCULO
# ─────────────────────────────────────────────────────────────────
def mapear_strings_a_filas(ws_arreglo) -> dict:
    """
    Recorre la hoja 'Arreglo mppts SM' y devuelve un dict:
    { inv_num: [(row, col_J, col_K), ...], ... }

    Detecta bloques de inversores buscando celdas que contengan
    "INVERSOR" en la columna B.  Dentro de cada bloque, recoge
    las filas donde la columna D tiene un valor numérico > 0
    (son las filas de strings activos).
    """
    MAX_ROW = ws_arreglo.max_row

    # Encontrar filas de encabezado de cada inversor
    inv_rows = {}
    for r in range(1, MAX_ROW + 1):
        val = ws_arreglo.cell(r, 2).value          # columna B
        if val and "INVERSOR" in str(val).upper():
            # Extraer número de inversor: "INVERSOR 1" → 1
            parts = str(val).upper().split()
            try:
                num = int(parts[parts.index("INVERSOR") + 1])
                inv_rows[num] = r
            except (ValueError, IndexError):
                pass

    if not inv_rows:
        raise RuntimeError("No se encontraron bloques de INVERSOR en la hoja 'Arreglo mppts SM'")

    inv_nums = sorted(inv_rows.keys())

    # Columna J = 10, columna K = 11 (1-indexed)
    COL_J = 10
    COL_K = 11
    COL_D = 4

    mapa = {}
    for i, inv_num in enumerate(inv_nums):
        start = inv_rows[inv_num] + 1          # fila siguiente al encabezado
        # El bloque termina justo antes del siguiente encabezado (o al final)
        if i + 1 < len(inv_nums):
            end = inv_rows[inv_nums[i + 1]] - 1
        else:
            end = MAX_ROW

        filas_strings = []
        for r in range(start, end + 1):
            d_val = ws_arreglo.cell(r, COL_D).value
            # Fila de string activo: D tiene un número > 0
            try:
                if d_val is not None and float(d_val) > 0:
                    filas_strings.append(r)
            except (ValueError, TypeError):
                pass

        mapa[inv_num] = [(r, COL_J, COL_K) for r in filas_strings]

    return mapa


# ─────────────────────────────────────────────────────────────────
# 4. ESCRIBIR EN EL EXCEL DE CÁLCULO
# ─────────────────────────────────────────────────────────────────
def llenar_excel(metrado: dict, proyecto: dict, excel_base: str, excel_salida: str):
    Path(excel_salida).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_base, keep_vba=True)

    # ── Hoja MEMORIA: datos del proyecto ──────────────────────────
    ws_mem = wb["MEMORIA"]
    ws_mem["C4"] = proyecto["nombre"]
    ws_mem["C5"] = proyecto["fecha"]
    ws_mem["C6"] = proyecto["ciudad"]
    ws_mem["C7"] = proyecto["departamento"]
    ws_mem["D7"] = proyecto["temp_prom"]
    ws_mem["E10"] = proyecto["temp_min"]
    ws_mem["E11"] = proyecto["temp_max"]

    # Longitudes cable AC por inversor (AB24, AB25, AB26, …)
    # AB33 (totalizador → transformador) se deja como está en el Excel base
    celdas_ac_inv = ["AB24", "AB25", "AB26", "AB27", "AB28"]
    for i, long_ac in enumerate(proyecto.get("long_ac_inv", [])):
        if i >= len(celdas_ac_inv):
            break
        ws_mem[celdas_ac_inv[i]] = long_ac

    print(f"[MEMORIA] Datos del proyecto escritos.")
    print(f"[MEMORIA] Longitudes AC inversores: {proyecto.get('long_ac_inv', [])}")

    # ── Hoja Arreglo mppts SM: longitudes de cable por string ─────
    ws_arr = wb["Arreglo mppts SM"]
    mapa   = mapear_strings_a_filas(ws_arr)

    # Organizar metrado por inversor: I1Sn → inv 1, I2Sn → inv 2, …
    por_inversor = {}
    for string_id, (pos, neg) in metrado.items():
        try:
            # Formato esperado: I{inv}S{str}
            partes = string_id.upper().lstrip("I").split("S")
            inv_num = int(partes[0])
            str_num = int(partes[1])
            por_inversor.setdefault(inv_num, {})[str_num] = (pos, neg)
        except (ValueError, IndexError):
            print(f"[ADVERTENCIA] String ID no reconocido: {string_id}")

    total_escritas = 0
    for inv_num, strings_inv in sorted(por_inversor.items()):
        filas = mapa.get(inv_num, [])
        strings_ordenados = sorted(strings_inv.items())  # ordenados por número de string

        if len(strings_ordenados) > len(filas):
            print(
                f"[ADVERTENCIA] Inversor {inv_num}: {len(strings_ordenados)} strings en metrado "
                f"pero solo {len(filas)} filas disponibles en el Excel."
            )

        for idx, (str_num, (pos_m, neg_m)) in enumerate(strings_ordenados):
            if idx >= len(filas):
                break
            row, col_j, col_k = filas[idx]
            ws_arr.cell(row, col_j).value = pos_m
            ws_arr.cell(row, col_k).value = neg_m
            total_escritas += 1

        print(f"[Arreglo] Inversor {inv_num}: {min(len(strings_ordenados), len(filas))} strings escritos.")

    # ── Guardar ───────────────────────────────────────────────────
    wb.save(excel_salida)
    print(f"\n✓ Excel guardado en:\n  {excel_salida}")
    print(f"  Total strings escritos: {total_escritas}")


# ─────────────────────────────────────────────────────────────────
# 5. MAPA DE TABLAS A CAPTURAR COMO IMAGEN
#
# Cada entrada define:
#   "hoja"        : nombre de la hoja en el Excel de cálculo
#   "rango"       : celda_inicio:celda_fin a capturar como imagen
#   "tabla"       : identificador interno
#   "ancla_word"  : texto parcial del caption en el Word donde se
#                   inserta la imagen JUSTO ANTES del caption.
#                   Se busca por título conceptual (no por número)
#                   para que funcione en cualquier operador de red.
# ─────────────────────────────────────────────────────────────────
TABLA_MAP = [
    # ── Regulación DC ── cols B:M, sin fila TOTAL ───────────────────
    {
        "hoja":       "Arreglo mppts SM",
        "rango":      "B2:M32",
        "tabla":      "reg_dc_inv1",
        "ancla_word": "regulación DC inversor 1",
    },
    {
        "hoja":       "Arreglo mppts SM",
        "rango":      "B36:M66",
        "tabla":      "reg_dc_inv2",
        "ancla_word": "regulación DC inversor 2",
    },
    {
        "hoja":       "Arreglo mppts SM",
        "rango":      "B70:M100",
        "tabla":      "reg_dc_inv3",
        "ancla_word": "regulación DC inversor 3",
    },
    # ── Pérdidas DC ── cols B:K ──────────────────────────────────────
    {
        "hoja":       "Pérdidas DC",
        "rango":      "B2:K34",
        "tabla":      "perd_dc_inv1",
        "ancla_word": "pérdidas DC inversor 1",
    },
    {
        "hoja":       "Pérdidas DC",
        "rango":      "B35:K67",
        "tabla":      "perd_dc_inv2",
        "ancla_word": "pérdidas DC inversor 2",
    },
    {
        "hoja":       "Pérdidas DC",
        "rango":      "B68:K103",
        "tabla":      "perd_dc_inv3",
        "ancla_word": "pérdidas DC inversor 3",
    },
    # ── AC ───────────────────────────────────────────────────────────
    {
        "hoja":       "Regulación AC",
        "rango":      "B2:M11",
        "tabla":      "reg_ac",
        "ancla_word": "regulación AC",
    },
    {
        "hoja":       "Pérdidas AC",
        "rango":      "B2:I10",
        "tabla":      "perd_ac",
        "ancla_word": "pérdidas de energía AC",
    },
]


# ─────────────────────────────────────────────────────────────────
# 6. MAIN
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== Auto-GD — Fase 1: Llenar Excel de cálculo ===\n")
    metrado = leer_metrado(EXCEL_METRADO)
    print(f"Metrado cargado: {len(metrado)} strings\n")
    llenar_excel(metrado, PROYECTO, EXCEL_BASE, EXCEL_SALIDA)
    print(f"\nTablas configuradas para captura (Fase 2): {len(TABLA_MAP)}")
    for t in TABLA_MAP:
        print(f"  [{t['tabla']}]  {t['hoja']}  {t['rango']}")
