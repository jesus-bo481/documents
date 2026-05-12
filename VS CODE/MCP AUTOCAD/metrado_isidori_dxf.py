"""
metrado_isidori_dxf.py
Metrado de cable DC para el DXF ISIDORI CLD TRAINING.

Diferencias vs MOMOTUS:
  - Conductores en "Cable Solar DC Strings 3" (todos color 18, sin distincion pos/neg)
  - Conductores horizontales; su ENDPOINT esta cerca de la etiqueta IxSy
  - Matching: por distancia 2D del endpoint al label (no por x_start)
  - Relleno con ceros: inversores con menos strings que MAX_STRINGS
    quedan con filas en cero hasta completar MAX_STRINGS por inversor

Output Excel = mismo formato que metrado_guiado_strings.lsp
"""
import re
import math
import sys
from pathlib import Path

try:
    import ezdxf
except ImportError:
    sys.exit("pip install ezdxf")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuracion ────────────────────────────────────────────────────────────
SOURCE          = r"C:\Users\JesúsAndrésBustilloO\Documents\ISIDORI CLD TRAINING.dxf"
OUTPUT_XLSX     = r"C:\Users\JesúsAndrésBustilloO\Documents\VS CODE\MCP AUTOCAD\output\metrado_ISIDORI.xlsx"
OUTPUT_CSV      = r"C:\Users\JesúsAndrésBustilloO\Documents\VS CODE\MCP AUTOCAD\output\metrado_ISIDORI.csv"

CONDUCTOR_LAYER = "Cable Solar DC Strings 3"
LABEL_LAYER     = "CONEXION"
NUM_INVERSORES  = 5
MAX_STRINGS     = 28   # rellenar con ceros hasta este numero por inversor
MATCH_TOL       = 5.0  # m — tolerancia para endpoint → label

IXSY_RE         = re.compile(r"I(\d+)S(\d+)")

# Paleta de colores por inversor (Excel)
INV_FILLS = ["DCE6F1", "EBF5EB", "FFF2CC", "F2DCDB", "E8E0F0", "FCE4D6"]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def poly_length(pts) -> float:
    total = 0.0
    for i in range(len(pts) - 1):
        dx = float(pts[i+1][0]) - float(pts[i][0])
        dy = float(pts[i+1][1]) - float(pts[i][1])
        total += math.hypot(dx, dy)
    return total


def parse_ixsy(raw: str):
    m = IXSY_RE.search(raw)
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def endpoint_of(pts) -> tuple:
    """Devuelve (x_end, y_end) del ultimo punto de la polyline."""
    return float(pts[-1][0]), float(pts[-1][1])


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"\n=== METRADO ISIDORI ===")
    print(f"  Fuente : {SOURCE}")

    doc = ezdxf.readfile(SOURCE)
    msp = doc.modelspace()

    # 1. Recolectar conductores con su endpoint y longitud
    conductors = []
    for e in msp:
        if e.dxftype() != "LWPOLYLINE" or e.dxf.layer != CONDUCTOR_LAYER:
            continue
        try:
            pts = list(e.get_points())
            if len(pts) < 2:
                continue
            xe, ye = endpoint_of(pts)
            conductors.append({
                "xe":    xe,
                "ye":    ye,
                "len":   poly_length(pts),
                "used":  False,
            })
        except Exception as ex:
            print(f"  WARN conductor: {ex}")
    print(f"  Conductores en '{CONDUCTOR_LAYER}': {len(conductors)}")

    # 2. Recolectar etiquetas IxSy
    labels = []
    for e in msp:
        if e.dxftype() != "MTEXT" or e.dxf.layer != LABEL_LAYER:
            continue
        try:
            inv, str_n = parse_ixsy(e.text)
            if inv is None:
                continue
            labels.append({
                "inv":   inv,
                "str_n": str_n,
                "x":     float(e.dxf.insert.x),
                "y":     float(e.dxf.insert.y),
                "tag":   f"I{inv}S{str_n}",
            })
        except Exception as ex:
            print(f"  WARN etiqueta: {ex}")
    print(f"  Etiquetas IxSy en '{LABEL_LAYER}': {len(labels)}")

    # Resumen por inversor
    from collections import Counter
    inv_cnt = Counter(l["inv"] for l in labels)
    for i in sorted(inv_cnt):
        print(f"    I{i}: {inv_cnt[i]} strings en DXF")

    # 3. Asociacion: para cada etiqueta, los 2 conductores con endpoint mas cercano
    #    Se consumen (used=True) para no reusar el mismo conductor en otra etiqueta.
    results = []   # (tag, inv, str_n, len_a, len_b)

    # Ordenar etiquetas para procesar de manera consistente (por inv, str_n)
    labels_sorted = sorted(labels, key=lambda l: (l["inv"], l["str_n"]))

    for lbl in labels_sorted:
        xl, yl = lbl["x"], lbl["y"]
        # Distancia 2D del endpoint de cada conductor no usado al label
        dists = []
        for ci, c in enumerate(conductors):
            if c["used"]:
                continue
            d = math.hypot(c["xe"] - xl, c["ye"] - yl)
            dists.append((d, ci))
        dists.sort()

        # Tomar los 2 mas cercanos dentro de tolerancia
        selected = []
        for d, ci in dists[:2]:
            if d <= MATCH_TOL:
                selected.append(ci)
                conductors[ci]["used"] = True

        if len(selected) == 2:
            ca = conductors[selected[0]]
            cb = conductors[selected[1]]
            # Asignar: el de mayor longitud = positivo (convencion arbitraria)
            len_pos = max(ca["len"], cb["len"])
            len_neg = min(ca["len"], cb["len"])
            results.append((lbl["tag"], lbl["inv"], lbl["str_n"], len_pos, len_neg))
        elif len(selected) == 1:
            ca = conductors[selected[0]]
            results.append((lbl["tag"], lbl["inv"], lbl["str_n"], ca["len"], 0.0))
            print(f"  WARN: {lbl['tag']} solo 1 conductor encontrado (dist={dists[0][0]:.2f}m)")
        else:
            results.append((lbl["tag"], lbl["inv"], lbl["str_n"], 0.0, 0.0))
            if dists:
                print(f"  WARN: {lbl['tag']} sin conductores en {MATCH_TOL}m (mejor dist={dists[0][0]:.2f}m)")
            else:
                print(f"  WARN: {lbl['tag']} sin conductores disponibles")

    used_count = sum(1 for c in conductors if c["used"])
    print(f"\n  Conductores asignados: {used_count} / {len(conductors)}")
    if used_count < len(conductors):
        print(f"  Conductores no asignados: {len(conductors) - used_count}")

    # 4. Rellenar con ceros hasta MAX_STRINGS por inversor
    results.sort(key=lambda r: (r[1], r[2]))

    final_results = []
    for inv in range(1, NUM_INVERSORES + 1):
        inv_results = [r for r in results if r[1] == inv]
        existing_strs = {r[2] for r in inv_results}
        inv_results_dict = {r[2]: r for r in inv_results}

        for s in range(1, MAX_STRINGS + 1):
            if s in inv_results_dict:
                final_results.append(inv_results_dict[s])
            else:
                # Rellenar con cero
                final_results.append((f"I{inv}S{s}", inv, s, 0.0, 0.0))

    # Verificar total
    print(f"\n  Filas en Excel: {len(final_results)} "
          f"({NUM_INVERSORES} inv x {MAX_STRINGS} strings)")

    # 5. Preview
    print(f"\n  {'STRING':8s}  {'POS(m)':>10s}  {'NEG(m)':>10s}  {'TOTAL(m)':>10s}")
    print(f"  {'-'*8}  {'-'*10}  {'-'*10}  {'-'*10}")
    for tag, inv, str_n, lp, ln in final_results[:14]:
        total = lp + ln
        flag  = "  <- CERO" if total == 0 else ""
        print(f"  {tag:8s}  {lp:10.3f}  {ln:10.3f}  {total:10.3f}{flag}")
    if len(final_results) > 14:
        print(f"  ... ({len(final_results)-14} mas)")

    # Totales
    total_pos = sum(r[3] for r in final_results)
    total_neg = sum(r[4] for r in final_results)
    print(f"\n  TOTAL POS : {total_pos:.2f} m")
    print(f"  TOTAL NEG : {total_neg:.2f} m")
    print(f"  TOTAL DC  : {total_pos + total_neg:.2f} m")

    # 6. Exportar
    export_xlsx(final_results)
    export_csv(final_results)


# ─── Exportacion ──────────────────────────────────────────────────────────────

def export_xlsx(results: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrado DC"

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezado
    headers = ["STRING", "POSITIVO (m)", "NEGATIVO (m)", "TOTAL (m)"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 18

    # Datos
    current_inv = None
    fill_idx    = -1
    zero_fill   = PatternFill("solid", fgColor="F2F2F2")   # gris para filas cero

    for row_n, (tag, inv, str_n, lp, ln) in enumerate(results, 2):
        if inv != current_inv:
            current_inv = inv
            fill_idx = (fill_idx + 1) % len(INV_FILLS)

        is_zero = (lp == 0.0 and ln == 0.0)
        row_fill = zero_fill if is_zero else PatternFill("solid", fgColor=INV_FILLS[fill_idx])

        vals = [tag, round(lp, 3), round(ln, 3), round(lp + ln, 3)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col, value=val)
            c.fill = row_fill; c.border = border
            c.alignment = Alignment(horizontal="center")
            if col > 1:
                c.number_format = "0.000"
                if is_zero:
                    c.font = Font(color="AAAAAA", italic=True)

    # Totales
    last_data = len(results) + 1
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    for col in [2, 3, 4]:
        col_l = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUMIF(A2:A{last_data},\"<>0\"&\"\",{col_l}2:{col_l}{last_data})")
        c.font = Font(bold=True); c.number_format = "0.000"; c.border = border
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    # Fila de totales reales (incluyendo todo)
    total_row2 = total_row + 1
    ws.cell(row=total_row2, column=1, value="TOTAL BRUTO").font = Font(bold=True, size=10, italic=True)
    for col in [2, 3, 4]:
        col_l = get_column_letter(col)
        c = ws.cell(row=total_row2, column=col,
                    value=f"=SUM({col_l}2:{col_l}{last_data})")
        c.font = Font(bold=True, italic=True, size=10)
        c.number_format = "0.000"; c.border = border
        c.alignment = Alignment(horizontal="center")

    # Columnas
    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16
    ws.freeze_panes = "A2"

    Path(OUTPUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\n  Excel : {OUTPUT_XLSX}")


def export_csv(results: list):
    lines = ["STRING,POSITIVO_M,NEGATIVO_M,TOTAL_M"]
    for tag, inv, str_n, lp, ln in results:
        lines.append(f"{tag},{lp:.3f},{ln:.3f},{lp+ln:.3f}")
    Path(OUTPUT_CSV).parent.mkdir(parents=True, exist_ok=True)
    Path(OUTPUT_CSV).write_text("\n".join(lines), encoding="utf-8")
    print(f"  CSV   : {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
