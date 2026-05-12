"""
metrado_strings_dxf.py
Asocia automáticamente los conductores DC (LWPOLYLINE en STRINGS_AUTO) con las
etiquetas IxSy (MTEXT en CONEXION) y calcula el metrado de cable por string.

Lógica de asociación:
  - Cada mesa tiene strings_per_mesa * 2 conductores que EMPIEZAN en
    x_start = mesa.x_right - 2*panel_w - MODULE_GAP  (borde 2do panel desde bajante)
  - Se agrupa por x_start → identifica la mesa dueña de cada conductor
  - Dentro de una mesa, los conductores se ordenan por y descendente:
      [red_top, blue_top, red_bot, blue_bot]  → strings [top_label, bot_label]
  - La etiqueta IxSy con mayor y en esa mesa → par (red_top, blue_top)
  - La etiqueta IxSy con menor y → par (red_bot, blue_bot)

Output Excel (idéntico al CSV de metrado_guiado_strings.lsp):
  STRING | POSITIVO_M | NEGATIVO_M | TOTAL_M
"""
import re
import math
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import ezdxf
except ImportError:
    sys.exit("pip install ezdxf")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from generate_strings_dxf import (
    collect_mesas, get_block_bbox, detect_layout,
    MODULE_GAP, LAYER_STRINGS, LAYER_IXS,
)

# ─── Configuración ────────────────────────────────────────────────────────────
SOURCE      = r"C:\Users\JesúsAndrésBustilloO\Documents\VS CODE\MCP AUTOCAD\output\MOMOTUS_strings_v6_final2.dxf"
OUTPUT_XLSX = r"C:\Users\JesúsAndrésBustilloO\Documents\VS CODE\MCP AUTOCAD\output\metrado_strings.xlsx"
OUTPUT_CSV  = r"C:\Users\JesúsAndrésBustilloO\Documents\VS CODE\MCP AUTOCAD\output\metrado_strings.csv"

PANEL_BLOCK  = "PANEL_615"
BAJANTE_SIDE = "R"
COLOR_POS    = 1   # rojo  = positivo
COLOR_NEG    = 5   # azul  = negativo
IXSY_RE      = re.compile(r"I(\d+)S(\d+)")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def poly_length(pts) -> float:
    """Longitud total de una LWPOLYLINE dada su lista de puntos (x,y,...)."""
    total = 0.0
    for i in range(len(pts) - 1):
        dx = float(pts[i+1][0]) - float(pts[i][0])
        dy = float(pts[i+1][1]) - float(pts[i][1])
        total += math.hypot(dx, dy)
    return total


def parse_ixsy(raw_text: str):
    """Extrae (inv, str_num) de '\\pxt10;I2S5'. Retorna (None, None) si no coincide."""
    m = IXSY_RE.search(raw_text)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def x_start_expected(mesa: dict, pw: float, bajante_side: str) -> float:
    """X del primer punto de los conductores de esta mesa (2do panel desde bajante)."""
    if bajante_side.upper() == "R":
        return mesa["x_right"] - 2.0 * pw - MODULE_GAP
    else:
        return mesa["x_left"] + 2.0 * pw + MODULE_GAP


def find_mesa_by_xstart_y(x: float, y: float, mesas: list, pw: float,
                          bajante_side: str) -> dict | None:
    """
    Devuelve la mesa cuyo x_start_expected ≈ x Y cuyo rango y contiene y_line.
    Primero filtra por x (tol = pw*0.55), luego elige la de y más cercana.
    """
    tol_x = pw * 0.55
    candidates = [m for m in mesas
                  if abs(x - x_start_expected(m, pw, bajante_side)) <= tol_x]
    if not candidates:
        return None
    # Entre los candidatos, el que tenga y_line más cercano a su centro vertical
    def y_dist(mesa):
        yc = (mesa["y_bottom"] + mesa["y_top"]) / 2.0
        return abs(y - yc)
    return min(candidates, key=y_dist)


def find_mesa_for_label(x: float, y: float, mesas: list,
                        margin_x: float = 2.0, margin_y: float = 2.0) -> dict | None:
    """Mesa que contiene o está más cerca del punto (x, y)."""
    best, best_dist = None, float("inf")
    for mesa in mesas:
        # Distancia al bbox con expansión marginal
        cx = max(mesa["x_left"] - margin_x, min(x, mesa["x_right"] + margin_x))
        cy = max(mesa["y_bottom"] - margin_y, min(y, mesa["y_top"] + margin_y))
        d  = math.hypot(x - cx, y - cy)
        if d < best_dist:
            best_dist, best = d, mesa
    tol = max(margin_x, margin_y)
    return best if best_dist <= tol else None


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"\n=== METRADO STRINGS DXF ===")
    print(f"  Fuente : {SOURCE}")

    doc = ezdxf.readfile(SOURCE)
    msp = doc.modelspace()

    # 1. Layout
    info = detect_layout(doc)
    bbox = get_block_bbox(doc, PANEL_BLOCK)
    if bbox is None:
        sys.exit(f"ERROR: bloque '{PANEL_BLOCK}' no encontrado")

    bx0, by0, bx1, by1 = bbox
    pw          = bx1 - bx0
    strings_pm  = info["strings_per_mesa"]
    print(f"  Bloque : {PANEL_BLOCK}  pw={pw:.4f} m  strings/mesa={strings_pm}")

    mesas = collect_mesas(doc, PANEL_BLOCK, "individual_panels",
                          panel_rows_per_mesa=strings_pm, skip_layers={"0"})
    print(f"  Mesas  : {len(mesas)}")
    if not mesas:
        sys.exit("ERROR: no se encontraron mesas.")

    # 2. Recolectar conductores (LWPOLYLINE en STRINGS_AUTO)
    conductors = []  # {'x_start', 'y_line', 'color', 'length'}
    for e in msp:
        if e.dxftype() != "LWPOLYLINE" or e.dxf.layer != LAYER_STRINGS:
            continue
        try:
            pts = list(e.get_points())
            if len(pts) < 2:
                continue
            conductors.append({
                "x_start": float(pts[0][0]),
                "y_line":  float(pts[0][1]),
                "color":   int(e.dxf.color),
                "length":  poly_length(pts),
            })
        except Exception as ex:
            print(f"  WARN conductor: {ex}")
    print(f"  Conductores en {LAYER_STRINGS}: {len(conductors)}")

    # 3. Recolectar etiquetas IxSy (MTEXT en CONEXION)
    labels = []  # {'inv', 'str_n', 'x', 'y', 'tag'}
    for e in msp:
        if e.dxftype() != "MTEXT" or e.dxf.layer != LAYER_IXS:
            continue
        try:
            inv, str_n = parse_ixsy(e.text)
            if inv is None:
                continue
            labels.append({
                "inv":   inv,
                "str_n": str_n,
                "x":     float(e.dxf.insert.x),
                "y":     float(e.dxf.insert.y),
                "tag":   f"I{inv}S{str_n}",
            })
        except Exception as ex:
            print(f"  WARN etiqueta: {ex}")
    print(f"  Etiquetas IxSy en {LAYER_IXS}: {len(labels)}")

    # 4. Agrupar conductores por mesa (key = (x_left, y_bottom) redondeado)
    cond_by_mesa: dict[tuple, list] = {}
    unmatched_cond = 0
    for c in conductors:
        mesa = find_mesa_by_xstart_y(c["x_start"], c["y_line"], mesas, pw, BAJANTE_SIDE)
        if mesa is None:
            unmatched_cond += 1
            continue
        key = (round(mesa["x_left"], 2), round(mesa["y_bottom"], 2))
        cond_by_mesa.setdefault(key, []).append(c)

    if unmatched_cond:
        print(f"  Conductores sin mesa asignada : {unmatched_cond}")

    # 5. Agrupar etiquetas por mesa
    lbl_by_mesa: dict[tuple, list] = {}
    unmatched_lbl = 0
    for lbl in labels:
        mesa = find_mesa_for_label(lbl["x"], lbl["y"], mesas)
        if mesa is None:
            unmatched_lbl += 1
            print(f"  WARN: {lbl['tag']} sin mesa ({lbl['x']:.1f},{lbl['y']:.1f})")
            continue
        key = (round(mesa["x_left"], 2), round(mesa["y_bottom"], 2))
        lbl_by_mesa.setdefault(key, []).append(lbl)

    if unmatched_lbl:
        print(f"  Etiquetas sin mesa asignada  : {unmatched_lbl}")

    # 6. Asociar por mesa
    #    Conductores ordenados por y desc: [red_top, blue_top, red_bot, blue_bot]
    #    Etiquetas   ordenadas por y desc: [top_label, bot_label]
    #    Par i → conductores [i*2, i*2+1]
    results = []   # (tag, inv, str_n, len_pos, len_neg)
    warnings = []

    for key, lbls in lbl_by_mesa.items():
        conds = cond_by_mesa.get(key, [])

        # Ordenar por y descendente
        lbls_sorted  = sorted(lbls,  key=lambda l: -l["y"])
        conds_sorted = sorted(conds, key=lambda c: -c["y_line"])

        n_expected = strings_pm  # = len(lbls_sorted) nominalmente
        if len(conds_sorted) < n_expected * 2:
            warnings.append(
                f"  WARN mesa {key}: {len(conds_sorted)} conductores "
                f"para {n_expected} strings ({len(lbls_sorted)} etiquetas)"
            )
            for lbl in lbls_sorted:
                results.append((lbl["tag"], lbl["inv"], lbl["str_n"], 0.0, 0.0))
            continue

        for i, lbl in enumerate(lbls_sorted):
            pair = conds_sorted[i*2 : i*2+2]
            pos_cond = next((c for c in pair if c["color"] == COLOR_POS), None)
            neg_cond = next((c for c in pair if c["color"] == COLOR_NEG), None)
            # Fallback si los colores no coinciden exactamente
            len_pos = pos_cond["length"] if pos_cond else (pair[0]["length"] if pair else 0.0)
            len_neg = neg_cond["length"] if neg_cond else (pair[1]["length"] if len(pair) > 1 else 0.0)
            results.append((lbl["tag"], lbl["inv"], lbl["str_n"], len_pos, len_neg))

    # Ordenar: por inversor, luego string
    results.sort(key=lambda r: (r[1], r[2]))

    print(f"\n  Strings con metrado: {len(results)}")
    if warnings:
        for w in warnings:
            print(w)

    # 7. Mostrar resumen
    print(f"\n  {'STRING':8s}  {'POS(m)':>10s}  {'NEG(m)':>10s}  {'TOTAL(m)':>10s}")
    print(f"  {'-'*8}  {'-'*10}  {'-'*10}  {'-'*10}")
    for tag, inv, str_n, lp, ln in results[:10]:   # primeros 10 como preview
        print(f"  {tag:8s}  {lp:10.2f}  {ln:10.2f}  {lp+ln:10.2f}")
    if len(results) > 10:
        print(f"  ... ({len(results)-10} más)")

    total_pos = sum(r[3] for r in results)
    total_neg = sum(r[4] for r in results)
    print(f"\n  TOTAL POS : {total_pos:.2f} m")
    print(f"  TOTAL NEG : {total_neg:.2f} m")
    print(f"  TOTAL     : {total_pos + total_neg:.2f} m")

    # 8. Exportar
    export_xlsx(results)
    export_csv(results)


# ─── Exportación ──────────────────────────────────────────────────────────────

# Paleta por inversor (fondo de fila)
INV_FILLS = ["DCE6F1", "EBF5EB", "FFF2CC", "F2DCDB", "E8E0F0", "FCE4D6"]


def export_xlsx(results: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrado DC"

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Encabezado ────────────────────────────────────────────────────────────
    headers = ["STRING", "POSITIVO (m)", "NEGATIVO (m)", "TOTAL (m)"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 18

    # ── Datos ─────────────────────────────────────────────────────────────────
    current_inv = None
    fill_idx    = -1
    for row_n, (tag, inv, str_n, lp, ln) in enumerate(results, 2):
        if inv != current_inv:
            current_inv = inv
            fill_idx = (fill_idx + 1) % len(INV_FILLS)
        row_fill = PatternFill("solid", fgColor=INV_FILLS[fill_idx])

        vals = [tag, round(lp, 3), round(ln, 3), round(lp + ln, 3)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=col, value=val)
            c.fill = row_fill; c.border = border
            c.alignment = Alignment(horizontal="center")
            if col > 1:
                c.number_format = "0.000"

    # ── Totales ───────────────────────────────────────────────────────────────
    last_data = len(results) + 1
    total_row = last_data + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    for col in [2, 3, 4]:
        col_l = get_column_letter(col)
        c = ws.cell(row=total_row, column=col,
                    value=f"=SUM({col_l}2:{col_l}{last_data})")
        c.font = Font(bold=True); c.number_format = "0.000"; c.border = border
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    # ── Formato columnas ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    Path(OUTPUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\n  Excel guardado : {OUTPUT_XLSX}")


def export_csv(results: list):
    lines = ["STRING,POSITIVO_M,NEGATIVO_M,TOTAL_M"]
    for tag, inv, str_n, lp, ln in results:
        lines.append(f"{tag},{lp:.3f},{ln:.3f},{lp+ln:.3f}")
    Path(OUTPUT_CSV).parent.mkdir(parents=True, exist_ok=True)
    Path(OUTPUT_CSV).write_text("\n".join(lines), encoding="utf-8")
    print(f"  CSV   guardado : {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
